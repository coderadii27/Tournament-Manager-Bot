"""Tournament control panel cog (?t)."""

from __future__ import annotations

import io

import discord
from discord.ext import commands

from state import get_guild, save_guild, update_guild

BRAND_COLOR = 0x9B5CF6
ACCENT_COLOR = 0x00E5FF

CATEGORY_NAME = "🏆 BRN ESPORTS"

CHANNEL_NAMES = [
    "—͟͞͞-⨳〢info",
    "—͟͞͞-⨳〢updates",
    "—͟͞͞-⨳〢rules",
    "—͟͞͞-⨳〢how-to-register",
    "—͟͞͞-⨳〢registration-format",
    "—͟͞͞-⨳〢registration",
    "—͟͞͞-⨳〢confirm-teams",
    "—͟͞͞-⨳〢roadmaps",
    "—͟͞͞-⨳〢schedule",
    "—͟͞͞-⨳〢point-table",
    "—͟͞͞-⨳〢query",
]

LOCKED_KEYWORDS = (
    "info", "updates", "rules", "how-to-register", "registration-format",
    "confirm-teams", "roadmaps", "schedule", "point-table",
)
OPEN_KEYWORDS = ("registration", "query")

REGISTRATION_FORMAT = """```
Team Name -

@player1   (IGL — first tag gets IDP role)
@player2
@player3
@player4
@player5
```
*Just tag your players in order. The number of tags must equal the team size.*"""


# ──────────────────────────────────────────────────────────────────────────────
# Panel embed
# ──────────────────────────────────────────────────────────────────────────────

def panel_embed(g: dict, guild: discord.Guild | None = None) -> discord.Embed:
    if g.get("closed"):
        status = "🔴 Closed"
    elif g.get("paused"):
        status = "🟡 Paused"
    elif g.get("running"):
        status = "🟢 Live"
    else:
        status = "⚪ Idle"

    embed = discord.Embed(
        title="🏆  BRN ESPORTS  —  TOURNAMENT CONTROL",
        description=(
            "**Welcome to the official tournament control hub.**\n"
            "Manage your event end-to-end from a single panel below.\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "🛠️  *Create / edit tournament*\n"
            "📂  *Auto-build all 11 channels*\n"
            "▶️  *Go live (notify in registration)*\n"
            "⏸️  *Pause & lock everything*\n"
            "🔴  *Close & shut down for good*\n"
            "📊  *Group teams • 🎟️ Slot manager*\n"
            "📑  *Export team list to Excel*\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
        ),
        color=BRAND_COLOR,
    )
    embed.add_field(name="🏷️  Tournament", value=f"```{g.get('tournament_name', 'EliteQ-tourny')}```", inline=True)
    embed.add_field(name="🎟️  Slots", value=f"```{len(g.get('teams', []))}/{g.get('max_slots', 16)}```", inline=True)
    embed.add_field(name="👥  Team Size", value=f"```{g.get('team_size', 5)} players```", inline=True)
    embed.add_field(name="📡  Status", value=f"**{status}**", inline=True)
    embed.add_field(name="📊  Groups", value=f"```{len(g.get('groups', {}))}```", inline=True)
    embed.add_field(name="\u200b", value="\u200b", inline=True)
    if guild and guild.icon:
        embed.set_thumbnail(url=guild.icon.url)
    embed.set_footer(text="BRN ESPORTS OFFICIAL BOT  •  Made with 💜 by Cyclopso")
    return embed


# ──────────────────────────────────────────────────────────────────────────────
# Channel helpers
# ──────────────────────────────────────────────────────────────────────────────

def _is_locked(name: str) -> bool:
    return any(k in name for k in LOCKED_KEYWORDS)


def _is_open(name: str) -> bool:
    return any(k in name for k in OPEN_KEYWORDS)


def _tournament_channels(guild: discord.Guild) -> list[discord.TextChannel]:
    return [c for c in guild.text_channels if c.name in CHANNEL_NAMES]


async def ensure_category_and_channels(guild: discord.Guild) -> tuple[discord.CategoryChannel, list[discord.TextChannel], dict]:
    category = discord.utils.get(guild.categories, name=CATEGORY_NAME)
    if category is None:
        category = await guild.create_category(CATEGORY_NAME, reason="Tournament setup")

    ids = {"registration_channel_id": None, "confirm_channel_id": None, "format_channel_id": None}
    channels: list[discord.TextChannel] = []
    for name in CHANNEL_NAMES:
        ch = discord.utils.get(guild.text_channels, name=name)
        if ch is None:
            ch = await guild.create_text_channel(name, category=category)
        channels.append(ch)
        if "registration-format" in name:
            ids["format_channel_id"] = ch.id
        elif "registration" in name:
            ids["registration_channel_id"] = ch.id
        if "confirm-teams" in name:
            ids["confirm_channel_id"] = ch.id
    return category, channels, ids


async def apply_running_perms(guild: discord.Guild) -> None:
    everyone = guild.default_role
    for ch in _tournament_channels(guild):
        if _is_open(ch.name):
            ow = discord.PermissionOverwrite(view_channel=True, send_messages=True, read_message_history=True, add_reactions=True)
        elif _is_locked(ch.name):
            ow = discord.PermissionOverwrite(view_channel=True, send_messages=False, read_message_history=True, add_reactions=False)
        else:
            continue
        try:
            await ch.set_permissions(everyone, overwrite=ow, reason="Tournament running")
        except discord.Forbidden:
            pass


async def apply_paused_perms(guild: discord.Guild) -> None:
    everyone = guild.default_role
    ow = discord.PermissionOverwrite(view_channel=True, send_messages=False, read_message_history=True, add_reactions=False)
    for ch in _tournament_channels(guild):
        try:
            await ch.set_permissions(everyone, overwrite=ow, reason="Tournament paused")
        except discord.Forbidden:
            pass


async def post_to_all_tournament_channels(guild: discord.Guild, text: str) -> int:
    sent = 0
    for ch in _tournament_channels(guild):
        try:
            await ch.send(text)
            sent += 1
        except Exception:
            pass
    return sent


async def post_to_registration(guild: discord.Guild, text: str) -> bool:
    g = get_guild(guild.id)
    rid = g.get("registration_channel_id")
    ch = guild.get_channel(rid) if rid else None
    if ch is None:
        ch = discord.utils.get(guild.text_channels, name="—͟͞͞-⨳〢registration")
    if ch is None:
        return False
    try:
        await ch.send(text)
        return True
    except Exception:
        return False


async def post_format_message(guild: discord.Guild) -> None:
    fch = discord.utils.get(guild.text_channels, name="—͟͞͞-⨳〢registration-format")
    if fch is None:
        return
    async for msg in fch.history(limit=20):
        if msg.author == guild.me and "Team Name" in (msg.content or ""):
            return
    embed = discord.Embed(
        title="📋 Registration Format",
        description=(
            "Send your registration in the registration channel using this format.\n"
            "Just tag your players. The **first tag is the IGL** and gets the IDP role automatically."
        ),
        color=BRAND_COLOR,
    )
    embed.set_footer(text="BRN ESPORTS OFFICIAL BOT")
    await fch.send(embed=embed)
    await fch.send(REGISTRATION_FORMAT)


async def delete_previous_panels(channel: discord.abc.Messageable, me: discord.ClientUser) -> None:
    """Remove any previous tournament panel messages this bot posted in the channel."""
    try:
        async for msg in channel.history(limit=30):
            if msg.author.id != me.id:
                continue
            if not msg.embeds:
                continue
            title = (msg.embeds[0].title or "")
            if "TOURNAMENT CONTROL" in title.upper():
                try:
                    await msg.delete()
                except Exception:
                    pass
    except Exception:
        pass


# ──────────────────────────────────────────────────────────────────────────────
# Modals
# ──────────────────────────────────────────────────────────────────────────────

class CreateTournamentModal(discord.ui.Modal, title="Create Tournament"):
    name = discord.ui.TextInput(label="Tournament Name", default="EliteQ-tourny", max_length=64)
    team_size = discord.ui.TextInput(label="Team Size (players per team)", default="5", max_length=2)
    slots = discord.ui.TextInput(label="Total Slots (e.g. 16, 32, 64)", default="16", max_length=3)

    async def on_submit(self, interaction: discord.Interaction):
        try:
            ts = int(str(self.team_size))
            sl = int(str(self.slots))
            assert 1 <= ts <= 10 and 2 <= sl <= 256
        except Exception:
            await interaction.response.send_message("Invalid numbers provided.", ephemeral=True)
            return
        update_guild(
            interaction.guild_id,
            {
                "tournament_name": str(self.name),
                "team_size": ts,
                "max_slots": sl,
                "running": False,
                "paused": False,
                "closed": False,
                "teams": [],
                "groups": {},
            },
        )
        g = get_guild(interaction.guild_id)
        embed = discord.Embed(
            title="✅ Tournament Created",
            description=f"**{g['tournament_name']}** is ready.\nTeam Size: `{g['team_size']}` • Slots: `{g['max_slots']}`",
            color=ACCENT_COLOR,
        )
        await interaction.response.send_message(embed=embed, ephemeral=True)


class EditSettingsModal(discord.ui.Modal, title="Edit Tournament Settings"):
    def __init__(self, current: dict):
        super().__init__(timeout=300)
        self.name = discord.ui.TextInput(
            label="Tournament Name", default=str(current.get("tournament_name", "EliteQ-tourny")),
            max_length=64,
        )
        self.team_size = discord.ui.TextInput(
            label="Team Size (players per team)", default=str(current.get("team_size", 5)),
            max_length=2,
        )
        self.slots = discord.ui.TextInput(
            label="Total Slots", default=str(current.get("max_slots", 16)), max_length=3,
        )
        self.add_item(self.name)
        self.add_item(self.team_size)
        self.add_item(self.slots)

    async def on_submit(self, interaction: discord.Interaction):
        try:
            ts = int(str(self.team_size))
            sl = int(str(self.slots))
            assert 1 <= ts <= 10 and 2 <= sl <= 256
        except Exception:
            await interaction.response.send_message("Invalid numbers provided.", ephemeral=True)
            return
        update_guild(
            interaction.guild_id,
            {
                "tournament_name": str(self.name),
                "team_size": ts,
                "max_slots": sl,
            },
        )
        g = get_guild(interaction.guild_id)
        embed = discord.Embed(
            title="✏️ Settings Updated",
            description=(
                f"**{g['tournament_name']}**\n"
                f"Team Size: `{g['team_size']}` • Slots: `{g['max_slots']}`"
            ),
            color=ACCENT_COLOR,
        )
        await interaction.response.send_message(embed=embed, ephemeral=True)


class ManageGroupsModal(discord.ui.Modal, title="Manage Groups"):
    group_count = discord.ui.TextInput(label="Number of Groups", default="2", max_length=2)

    async def on_submit(self, interaction: discord.Interaction):
        try:
            n = int(str(self.group_count))
            assert 1 <= n <= 32
        except Exception:
            await interaction.response.send_message("Invalid group count.", ephemeral=True)
            return
        g = get_guild(interaction.guild_id)
        teams = list(g.get("teams", []))
        groups: dict[str, list] = {chr(ord("A") + i): [] for i in range(n)}
        for idx, t in enumerate(teams):
            key = chr(ord("A") + (idx % n))
            groups[key].append(t.get("name", f"Team {idx + 1}"))
        g["groups"] = groups
        save_guild(interaction.guild_id, g)
        embed = discord.Embed(title="📊 Groups Distributed", color=BRAND_COLOR)
        for key, members in groups.items():
            embed.add_field(name=f"Group {key}", value="\n".join(f"• {m}" for m in members) if members else "*empty*", inline=True)
        embed.set_footer(text="BRN ESPORTS OFFICIAL BOT")
        await interaction.response.send_message(embed=embed, ephemeral=True)


# ──────────────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────────────

def build_team_excel(g: dict, guild: discord.Guild) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Teams"

    header = ["Slot", "Team Name", "IGL", "IGL Discord ID", "Players", "Player IDs"]
    ws.append(header)
    head_fill = PatternFill("solid", fgColor="9B5CF6")
    head_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    teams = g.get("teams", [])
    for i, t in enumerate(teams, start=1):
        igl_id = t.get("captain_id")
        igl_member = guild.get_member(int(igl_id)) if igl_id else None
        igl_name = igl_member.display_name if igl_member else (t.get("player_names") or ["?"])[0]
        players = ", ".join(t.get("player_names") or [])
        pids = ", ".join(str(p) for p in (t.get("player_ids") or []))
        ws.append([i, t.get("name", ""), igl_name, str(igl_id or ""), players, pids])

    # Column widths
    widths = [6, 28, 22, 22, 60, 60]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Meta sheet
    meta = wb.create_sheet("Tournament")
    meta.append(["Tournament", g.get("tournament_name", "")])
    meta.append(["Team Size", g.get("team_size", "")])
    meta.append(["Total Slots", g.get("max_slots", "")])
    meta.append(["Filled Slots", len(teams)])
    meta.append(["Status", "Closed" if g.get("closed") else ("Paused" if g.get("paused") else ("Running" if g.get("running") else "Idle"))])
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 40
    for row in range(1, 6):
        meta.cell(row=row, column=1).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────────────────────────────
# View
# ──────────────────────────────────────────────────────────────────────────────

class TournamentPanelView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(label="Create Tournament", style=discord.ButtonStyle.success, emoji="🛠️", custom_id="t:create", row=0)
    async def create_btn(self, interaction: discord.Interaction, _b):
        if not interaction.user.guild_permissions.manage_guild:
            await interaction.response.send_message("You need **Manage Server** permission.", ephemeral=True)
            return
        await interaction.response.send_modal(CreateTournamentModal())

    @discord.ui.button(label="Edit Settings", style=discord.ButtonStyle.secondary, emoji="✏️", custom_id="t:edit", row=0)
    async def edit_btn(self, interaction: discord.Interaction, _b):
        if not interaction.user.guild_permissions.manage_guild:
            await interaction.response.send_message("You need **Manage Server** permission.", ephemeral=True)
            return
        g = get_guild(interaction.guild_id)
        await interaction.response.send_modal(EditSettingsModal(g))

    @discord.ui.button(label="Create Channels", style=discord.ButtonStyle.primary, emoji="📂", custom_id="t:channels", row=0)
    async def channels_btn(self, interaction: discord.Interaction, _b):
        if not interaction.user.guild_permissions.manage_channels:
            await interaction.response.send_message("You need **Manage Channels** permission.", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True, thinking=True)
        guild = interaction.guild
        category, channels, ids = await ensure_category_and_channels(guild)
        await post_format_message(guild)
        await apply_running_perms(guild)
        update_guild(interaction.guild_id, {
            "registration_channel_id": ids["registration_channel_id"],
            "confirm_channel_id": ids["confirm_channel_id"],
        })
        embed = discord.Embed(
            title="📂 Channels Ready",
            description=f"Created/verified **{len(channels)}** channels under **{category.name}** and applied tournament permissions.",
            color=ACCENT_COLOR,
        )
        await interaction.followup.send(embed=embed, ephemeral=True)

    @discord.ui.button(label="Start Tournament", style=discord.ButtonStyle.success, emoji="▶️", custom_id="t:start", row=1)
    async def start_btn(self, interaction: discord.Interaction, _b):
        if not interaction.user.guild_permissions.manage_guild:
            await interaction.response.send_message("You need **Manage Server** permission.", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True, thinking=True)
        guild = interaction.guild
        category, channels, ids = await ensure_category_and_channels(guild)
        await post_format_message(guild)
        await apply_running_perms(guild)
        update_guild(interaction.guild_id, {
            "running": True, "paused": False, "closed": False,
            "registration_channel_id": ids["registration_channel_id"],
            "confirm_channel_id": ids["confirm_channel_id"],
        })
        g = get_guild(interaction.guild_id)
        text = (
            f"```TOURNAMENT IS LIVE```\n"
            f"**{g.get('tournament_name', 'Tournament')}** is now **OPEN** for registrations.\n"
            f"**Drop your filled registration in this channel.**"
        )
        ok = await post_to_registration(guild, text)
        await interaction.followup.send(
            f"▶️ Tournament started. Registration channel notified: **{'✅' if ok else '⚠️ failed'}**.",
            ephemeral=True,
        )

    @discord.ui.button(label="Pause Tournament", style=discord.ButtonStyle.secondary, emoji="⏸️", custom_id="t:pause", row=1)
    async def pause_btn(self, interaction: discord.Interaction, _b):
        if not interaction.user.guild_permissions.manage_guild:
            await interaction.response.send_message("You need **Manage Server** permission.", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True, thinking=True)
        guild = interaction.guild
        g = get_guild(interaction.guild_id)
        new_paused = not g.get("paused", False)
        update_guild(interaction.guild_id, {"paused": new_paused, "closed": False})

        if new_paused:
            await apply_paused_perms(guild)
            text = (
                f"```TOURNAMENT HAS PAUSED```\n"
                f"**{g.get('tournament_name', 'Tournament')}** is now paused.\n"
                f"**All channels are locked. You can only view, not send messages.**"
            )
            sent = await post_to_all_tournament_channels(guild, text)
            await interaction.followup.send(
                f"⏸️ Tournament paused. All **{sent}** channels locked & notified.", ephemeral=True
            )
        else:
            await apply_running_perms(guild)
            text = (
                f"```TOURNAMENT RESUMED```\n"
                f"**{g.get('tournament_name', 'Tournament')}** is back live.\n"
                f"**Channels have been unlocked. Game on.**"
            )
            sent = await post_to_all_tournament_channels(guild, text)
            await interaction.followup.send(
                f"▶️ Tournament resumed. **{sent}** channels unlocked & announced.", ephemeral=True
            )

    @discord.ui.button(label="Close Tournament", style=discord.ButtonStyle.danger, emoji="🔴", custom_id="t:close", row=1)
    async def close_btn(self, interaction: discord.Interaction, _b):
        if not interaction.user.guild_permissions.manage_guild:
            await interaction.response.send_message("You need **Manage Server** permission.", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True, thinking=True)
        guild = interaction.guild
        g = get_guild(interaction.guild_id)
        update_guild(interaction.guild_id, {"closed": True, "running": False, "paused": True})
        await apply_paused_perms(guild)
        text = (
            f"```TOURNAMENT CLOSED```\n"
            f"**{g.get('tournament_name', 'Tournament')}** has been officially closed.\n"
            f"**All channels are now permanently locked. Thanks to every team.**"
        )
        sent = await post_to_all_tournament_channels(guild, text)
        await interaction.followup.send(
            f"🔴 Tournament closed. All **{sent}** channels locked & notified.", ephemeral=True
        )

    @discord.ui.button(label="Manage Groups", style=discord.ButtonStyle.primary, emoji="📊", custom_id="t:groups", row=2)
    async def groups_btn(self, interaction: discord.Interaction, _b):
        if not interaction.user.guild_permissions.manage_guild:
            await interaction.response.send_message("You need **Manage Server** permission.", ephemeral=True)
            return
        await interaction.response.send_modal(ManageGroupsModal())

    @discord.ui.button(label="Slot Manager", style=discord.ButtonStyle.danger, emoji="🎟️", custom_id="t:slots", row=2)
    async def slots_btn(self, interaction: discord.Interaction, _b):
        if not interaction.user.guild_permissions.manage_channels:
            await interaction.response.send_message("You need **Manage Channels** permission.", ephemeral=True)
            return
        cog = interaction.client.get_cog("SlotManager")
        if cog is None:
            await interaction.response.send_message("Slot manager cog not loaded.", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True, thinking=True)
        admin_ch = await cog.setup_panel(interaction.guild, invoker=interaction.user)
        public_ch = await cog.setup_public_claim_channel(interaction.guild)
        await interaction.followup.send(
            f"🔒 Private admin slot manager: {admin_ch.mention}\n"
            f"📣 Public slot-cancel-claim channel: {public_ch.mention}",
            ephemeral=True,
        )

    @discord.ui.button(label="Export Excel", style=discord.ButtonStyle.success, emoji="📑", custom_id="t:excel", row=2)
    async def excel_btn(self, interaction: discord.Interaction, _b):
        if not interaction.user.guild_permissions.manage_guild:
            await interaction.response.send_message("You need **Manage Server** permission.", ephemeral=True)
            return
        await interaction.response.defer(ephemeral=True, thinking=True)
        g = get_guild(interaction.guild_id)
        teams = g.get("teams", [])
        if not teams:
            await interaction.followup.send("No teams registered yet.", ephemeral=True)
            return
        try:
            buf = build_team_excel(g, interaction.guild)
        except Exception as e:
            await interaction.followup.send(f"❌ Failed to build Excel: `{e}`", ephemeral=True)
            return
        safe = "".join(c if c.isalnum() else "_" for c in g.get("tournament_name", "tournament"))
        filename = f"{safe}_teams.xlsx"
        await interaction.followup.send(
            content=f"📑 Confirmed teams export — **{len(teams)}** teams.",
            file=discord.File(buf, filename=filename),
            ephemeral=True,
        )


class Tournament(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot
        self.bot.add_view(TournamentPanelView())

    @commands.command(name="t")
    async def tournament_panel(self, ctx: commands.Context):
        if ctx.guild is None:
            await ctx.reply("Use this in a server.", mention_author=False)
            return
        await delete_previous_panels(ctx.channel, self.bot.user)
        try:
            await ctx.message.delete()
        except Exception:
            pass
        g = get_guild(ctx.guild.id)
        await ctx.send(embed=panel_embed(g, ctx.guild), view=TournamentPanelView())


async def setup(bot: commands.Bot):
    await bot.add_cog(Tournament(bot))
